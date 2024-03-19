from docxtpl import DocxTemplate
from typing import Union
import pandas as pd
from openpyxl import load_workbook
import re

__all__ = [
    "fill_docx",
    "fill_xlsx",
]

def fill_docx(
        data: Union[pd.Series, dict],
        template: str,
        full_path:str,   
):
    """
    Fill elements from 'data' into a docx template file, and save to the path specified by 'full_path'.
    :param data: Data to fill, can be of type pandas.Series or dict
    :param template: Path to the template file for data filling
    :param full_path: Path where the filled file is saved
    
    """

    # Load the template file
    doc = DocxTemplate(template)

    # If the provided data is of type pd.Series, convert it to dict
    if isinstance(data, pd.Series):
        data = data.to_dict()

    # Using the render method of docxtpl to fill the template
    doc.render(data)

    # Save to the specified path
    doc.save(full_path)

def fill_xlsx(
    data: Union[pd.Series, dict],
    template: str,
    full_path: str   
):
    """
    Fills elements from 'data' into a xlsx template file, and saves to the path specified by 'full_path'.
    :param data: Data to fill, can be of type pandas.Series or dict
    :param template: Path to the template file for data filling
    :param full_path: Path where the filled file will be saved
    """

    book = load_workbook(template)    # Load the template file

    # Iterate through all worksheets
    for sheet in book.worksheets:

        # Iterate through all cells in the worksheet
        for row in sheet.iter_rows():
            for cell in row:
                # If the cell value is a string, and contains a placeholder (e.g., '{{A}}')
                if isinstance(cell.value, str) and '{{' in cell.value:
                    # Extract the placeholder (e.g., 'A')
                    key = cell.value.strip('{}')
                    # If the placeholder can be found in the data, replace the placeholder with the corresponding data
                    if key in data:     
                        if is_only_placeholder(cell.value):
                            print(f'**{cell.value}**')
                            cell.value = data[key] 
                        else:          
                            placeholder =  '{{' + key + '}}'
                            cell_value = str(cell.value)  
                            print(f'**{cell_value}**')      
                            cell.value = cell_value.replace(placeholder, str(data[key]))

    # Save the result file
    book.save(full_path)




def is_only_placeholder(s: str) -> bool:
    """
    Check if a string s is a placeholder in the form of {{A}}
    :param s: The string to confirm if it's a placeholder 

    >>> is_placeholder('{{AA}}')
    True
    >>> is_placeholder('{{编号}}')
    True
    >>> is_placeholder('{{a_001}}')
    True
    >>> is_placeholder('{{AA}}BC')
    False
    """
    pattern = r'^\{\{[\w\u4e00-\u9fa5]+\}\}$'    # This pattern matches strings like {{A}} exactly 
    match = re.fullmatch(pattern, s)
    return match is not None        # If there is a match, the string is a placeholder, return True, otherwise, False

if __name__=="__main__":
    # import doctest
    # doctest.testmod()

    data = {
        "标题": "测试用标题", 
        "编号": 1,
        "姓名": "张三", 
        "性别": "男",
        "年龄": 15
    }
    fill_xlsx(data, template="d:/dev/filler/src/test/模板.xlsx", full_path="d:/dev/filler/src/test/out.xlsx")