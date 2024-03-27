from docxtpl import DocxTemplate
from typing import Union
import pandas as pd
from openpyxl import load_workbook
import re
from ._utils import is_empty

__all__ = [
    "fill_docx",
    "fill_xlsx",
]


def fill_docx(
    data: Union[pd.Series, dict],
    template: str,
    full_path: str,
):
    """
    Fill elements from 'data' into a docx template file, and save to the path specified by 'full_path'.
    :param data: Data to fill, can be of type pandas.Series or dict
    :param template: Path to the template file for data filling
    :param full_path: Path where the filled file is saved

    """

    # Load the template file
    doc = DocxTemplate(template)

    # If the provided data is of type pd.Series, convert it to dict
    if isinstance(data, pd.Series):
        data = data.to_dict()

    # Using the render method of docxtpl to fill the template
    doc.render(data)

    # Save to the specified path
    doc.save(full_path)


def fill_xlsx(data: Union[pd.Series, dict], template: str, full_path: str):
    """
    Fills elements from 'data' into a xlsx template file, and saves to the path specified by 'full_path'.
    :param data: Data to fill, can be of type pandas.Series or dict
    :param template: Path to the template file for data filling
    :param full_path: Path where the filled file will be saved
    """

    book = load_workbook(template)  # Load the template file

    # Iterate through all worksheets
    for sheet in book.worksheets:

        # Iterate through all cells in the worksheet
        for row in sheet.iter_rows():
            for cell in row:
                # If the cell value is a string, and contains a placeholder (e.g., '{{A}}')
                if isinstance(cell.value, str) and "{{" in cell.value:
                    # Extract the placeholder (e.g., 'A')
                    key = remove_after(remove_before(cell.value, "{{"), "}}").strip(
                        "{}"
                    )
                    # If the placeholder can be found in the data, replace the placeholder with the corresponding data
                    if key in data:
                        if is_only_placeholder(cell.value):
                            cell.value = data[key]
                        else:
                            placeholder = "{{" + key + "}}"
                            replaced = "" if is_empty(
                                data[key]) else str(data[key])
                            cell_value = str(cell.value)
                            cell.value = cell_value.replace(
                                placeholder, replaced)

    # Save the result file
    book.save(full_path)


def remove_before(s: str, spec: str):
    """
    Returns a new string with all characters removed before 'spec'
    :param s: The original string
    :param spec: The specific character
    """
    # Find the index of 'spec' in 's'
    idx = s.find(spec)

    # If 'spec' is not found in 's', return the original string
    if idx == -1:
        return s

    # If 'spec' is found, return the part of 's' from 'spec' to the end
    return s[idx:]


def remove_after(s: str, spec: str) -> str:
    """
    Return a new string removed all chars after spec
    :param s: The original string
    :param spec: The specific char
    """
    # Find the index of spec in s
    idx = s.find(spec)

    # If spec is not found in s, return the original string
    if idx == -1:
        return s

    # Otherwise, return the part of s before and including spec
    return s[: idx + len(spec)]


def is_only_placeholder(s: str) -> bool:
    """
    Check if a string s is a placeholder in the form of {{A}}
    :param s: The string to confirm if it's a placeholder

    >>> is_placeholder('{{AA}}')
    True
    >>> is_placeholder('{{编号}}')
    True
    >>> is_placeholder('{{a_001}}')
    True
    >>> is_placeholder('{{AA}}BC')
    False
    """
    pattern = r"^\{\{[\w\u4e00-\u9fa5]+\}\}$"  # This pattern matches strings like {{A}} exactly
    match = re.fullmatch(pattern, s)
    # If there is a match, the string is a placeholder, return True, otherwise, False
    return match is not None


if __name__ == "__main__":
    import doctest

    doctest.testmod()
